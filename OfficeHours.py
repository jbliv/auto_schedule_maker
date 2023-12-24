import openpyxl as xl
from openpyxl.styles import PatternFill
import copy


# Officer Class
class Officer:
    def __init__(self, name, schedule, order, reliable=1) -> None:
        self.__Name = name
        self.__Schedule = schedule
        self.__Order = order
        self.__Reliability = reliable
        self.__Preferred = 0

    # Getter Functions
    def getName(self) -> str:
        return self.__Name

    def getScheduleObject(self):
        return self.__Schedule

    def getSchedule(self) -> list:
        return self.__Schedule.getSchedule()

    def getOrder(self) -> int:
        return self.__Order

    def getReliability(self) -> int:
        return self.__Reliability

    # Availability Stats
    def ruleFollower(self) -> int:
        rulesFollowed = 0
        ruleCheck = self.__Schedule.availStats()
        if ruleCheck[3] <= 6:
            rulesFollowed += 1
        if ruleCheck[2] <= 12:
            rulesFollowed += 1
        return rulesFollowed

    def availableRatio(self) -> float:
        availabilityStats = self.__Schedule.availStats()
        return float(availabilityStats[1]) / float(availabilityStats[0])

    def __str__(self) -> str:
        return (
            "Name: "
            + self.__Name
            + " | Reliability: "
            + str(self.__Reliability)
            + " | Order: "
            + str(self.__Order)
            + " |\n"
            + self.__Schedule.__str__()
        )


# Schedule Class
class Schedule:
    def __init__(self, data) -> None:
        self.__Schedule = data

    def getSchedule(self) -> list:
        return self.__Schedule

    def availStats(self) -> list:
        unavailable = 0
        available = 0
        preferred = 0
        ideal = 0
        for hour in self.__Schedule:
            for day in hour:
                if day == 0:
                    unavailable += 1
                elif day == 1:
                    available += 1
                elif day == 2:
                    available += 1
                    preferred += 1
                elif day == 3:
                    available += 1
                    preferred += 1
                    ideal += 1
        return [unavailable, available, preferred, ideal]

    def hourSegments(self) -> list:
        data = self.getSchedule()
        output = [[[], [], [], [], []] for x in range(timeOpen)]
        for hour in range(timeOpen):
            for day in range(5):
                if data[2 * hour][day] <= data[2 * hour + 1][day]:
                    output[hour][day] = data[2 * hour][day]
                else:
                    output[hour][day] = data[2 * hour + 1][day]
        return Schedule(output)

    def idealOnly(self) -> list:
        output = []
        if self.availStats()[3] < 6:
            threeCheck = True
        else:
            threeCheck = False
        for hour in self.__Schedule:
            currentHour = []
            for day in hour:
                if day == 3:
                    currentHour.append(day)
                elif day == 2 and threeCheck:
                    currentHour.append(day)
                else:
                    currentHour.append(0)
            output.append(currentHour)
        return Schedule(output)

    def preferredOnly(self):
        output = []
        for hour in self.__Schedule:
            currentHour = []
            for day in hour:
                if day != 3 and day != 2:
                    currentHour.append(0)
                else:
                    currentHour.append(day)
            output.append(currentHour)
        return Schedule(output)

    def __str__(self) -> str:
        if len(self.__Schedule) == timeOpen:
            return "String for hour-segment schedules unavailable."
        hours = [
            "9:00 - 9:30  |",
            "9:30 - 10:00 |",
            "10:00 - 10:30|",
            "10:30 - 11:00|",
            "11:00 - 11:30|",
            "11:30 - 12:00|",
            "12:00 - 12:30|",
            "12:30 - 1:00 |",
            "1:00 - 1:30  |",
            "1:30 - 2:00  |",
            "2:00 - 2:30  |",
            "2:30 - 3:00  |",
            "3:00 - 3:30  |",
            "3:30 - 4:00  |",
            "4:00 - 4:30  |",
            "4:30 - 5:00  |",
        ]
        output = (
            "Hours:       | Monday      | Tuesday     | Wednesday   | Thursday    | Friday      |\n"
            + "____________________________________________________________________________________\n"
        )
        tracer = 0
        statsRaw = self.availStats()
        stats = []
        for i in statsRaw:
            stats.append(str(round((float(i) / 2), 1)))
        for time in self.__Schedule:
            currentHour = ""
            for day in time:
                if day == 0:
                    currentHour = currentHour + " Unavailable |"
                elif day == 1:
                    currentHour = currentHour + " Available   |"
                elif day == 2:
                    currentHour = currentHour + " Preferred   |"
                elif day == 3:
                    currentHour = currentHour + " Ideal       |"
            output = output + hours[tracer] + currentHour + "\n"
            tracer += 1
        output = (
            output
            + "Total Hrs Avail.: "
            + stats[1]
            + " | Total Hrs Unavail.: "
            + stats[0]
            + " | Pref. Hrs: "
            + stats[2]
            + " | Ideal Hrs "
            + stats[3]
            + " |"
        )
        return output


# Master Schedule Class
class MasterMaker:
    def __init__(self, officers) -> None:
        self.__Officers = [None] * len(officers)
        self.__NameDict = {}
        for officer in officers:
            self.__Officers[len(officers) - officer.getOrder()] = officer
            self.__NameDict[officer.getName()] = officer

    def nameAvailability(self, modOfficers) -> list:
        schedule = [[[], [], [], [], []] for x in range(timeOpen)]
        avail = [[[], [], [], [], []] for x in range(timeOpen)]
        pref = [[[], [], [], [], []] for x in range(timeOpen)]
        ideal = [[[], [], [], [], []] for x in range(timeOpen)]
        for officer in modOfficers:
            for time in range(len(officer.getSchedule())):
                for day in range(len(officer.getSchedule()[time])):
                    if officer.getSchedule()[time][day] == 1:
                        avail[time][day].append(officer.getName())
                    elif officer.getSchedule()[time][day] == 2:
                        pref[time][day].append(officer.getName())
                    elif officer.getSchedule()[time][day] == 3:
                        ideal[time][day].append(officer.getName())
        for time in range(len(schedule)):
            for day in range(len(schedule[time])):
                for officer in avail[time][day]:
                    schedule[time][day].append(officer)
                for officer in pref[time][day]:
                    schedule[time][day].append(officer)
                for officer in ideal[time][day]:
                    schedule[time][day].append(officer)
        return schedule

    def emptySlots(self, heatmap) -> bool:
        for time in heatmap:
            for day in time:
                if day == []:
                    return False
        return True

    def createSchedules(self, officers) -> list:
        # Initialize variables used in stack and set stack base case to variable base case
        currentSchedule = [[None, None, None, None, None] for x in range(timeOpen)]
        currentAvail = self.nameAvailability(officers)

        currentHoursDict = {i.getName(): 0 for i in officers}
        nameToObjectDict = {i.getName(): i for i in officers}
        currentUnidealSet = set()
        # Solution Comparison Variables
        minAvail = 40
        minPref = 40
        currentAvailCount = 0
        stack = [
            (
                currentSchedule,
                currentAvail,
                currentHoursDict,
                currentUnidealSet,
                currentAvailCount,
            )
        ]

        # Empty solutions set
        solutions = []

        # TESTING
        testing = 0
        testingHundo = 0

        # While loop that controls stack and current item from stack
        while stack != []:
            if len(stack) < 10000:
                stackCheck = 0
            else:
                stackCheck = -1
            testing += 1
            if testing == 100000:
                testingHundo += 1
                testing = 0
                print(testingHundo, " Hundred Thousand Nodes tested")
                print(len(stack), " items in Stack")
                print(len(solutions), " Solutions Found")
            # Boolean for tracking if an operation occured or branch killed on this loop,
            # if True no other operations occur on loop to prevent operations from interfering.
            operationDone = False

            # Variables for checking if a full schedule has been created
            filledSlots = 0
            fullSchedule = False

            # Sets schedule variables for instance of loop
            (
                currentSchedule,
                currentAvail,
                currentHoursDict,
                currentUnidealSet,
                currentAvailCount,
            ) = stack.pop(stackCheck)
            if currentAvailCount <= minAvail:
                # Check if branch is dead, complete, or contains slots with 1 possible option
                for time in range(len(currentAvail)):
                    for day in range(len(currentAvail[time])):
                        # Checks if operation has been done and kills loop if so
                        if operationDone:
                            break
                        # Checks if slot has been filled in schedule and stops further checks from being done on slot
                        elif currentSchedule[time][day] != None:
                            filledSlots += 1

                            # Checks if schedule is full
                            if filledSlots == 40:
                                # Stops further operations on loop and adds solution to schedule
                                operationDone = True
                                if currentAvailCount < minAvail:
                                    minAvail = currentAvailCount
                                    solutions = []
                                solutions.append((currentSchedule, currentHoursDict))

                        # Checks if slot has no possible availabilities
                        elif currentAvail[time][day] == []:
                            operationDone = True

                        # Checks if slot only has 1 possible availability.
                        elif len(currentAvail[time][day]) == 1:
                            operationDone = True

                            # Creates easier variable to call for lone officer
                            soleOfficer = currentAvail[time][day][0]

                            # Checks if hour is not preferred and then adds officer to unideal set if so or if already in unideal set kills branch
                            if (
                                nameToObjectDict[soleOfficer].getSchedule()[time][day]
                                == 1
                            ):
                                currentAvailCount += 1
                                if soleOfficer in currentUnidealSet:
                                    operationDone = True
                                    break
                                else:
                                    currentUnidealSet.add(soleOfficer)

                            # Makes change to current branch and checks if officer hits hours cap
                            currentSchedule[time][day] = soleOfficer
                            currentHoursDict[soleOfficer] += 1
                            if currentHoursDict[soleOfficer] >= 3:
                                for soleTime in currentAvail:
                                    for soleDay in soleTime:
                                        if soleOfficer in soleDay:
                                            soleDay.remove(soleOfficer)

                            # Ensure availability for that hour is empty and adds branch to stack
                            currentAvail[time][day] = []
                            stack.append(
                                (
                                    currentSchedule,
                                    currentAvail,
                                    currentHoursDict,
                                    currentUnidealSet,
                                    currentAvailCount,
                                )
                            )

                    # Checks if operation has been done and kills loop if so
                    if operationDone:
                        break
            else:
                operationDone = True

            # Checks if operationDone and if not adds all possible next branches to stack
            if not operationDone:
                for time in range(len(currentAvail)):
                    for day in range(len(currentAvail[time])):
                        for officer in currentAvail[time][day]:
                            # Creates copies of all variables that will be added to stack
                            operationDone = True
                            copySchedule = copy.deepcopy(currentSchedule)
                            copyAvail = copy.deepcopy(currentAvail)
                            copyHoursDict = copy.deepcopy(currentHoursDict)
                            copyUnidealSet = copy.deepcopy(currentUnidealSet)
                            copyAvailCount = copy.deepcopy(currentAvailCount)

                            # Boolean for checking if schedule change break 2 unideal rule
                            ruleFollowed = True

                            # Checks if 2 unideal hrs rule is broken and performs necessary actions
                            if nameToObjectDict[officer].getSchedule()[time][day] == 1:
                                copyAvailCount += 1
                                if officer in copyUnidealSet:
                                    ruleFollowed = False
                                else:
                                    copyUnidealSet.add(officer)

                            # Makes necessary changes
                            copySchedule[time][day] = officer
                            copyHoursDict[officer] += 1
                            if copyHoursDict[officer] >= 3:
                                for soleTime in copyAvail:
                                    for soleDay in soleTime:
                                        if officer in soleDay:
                                            soleDay.remove(officer)
                            copyAvail[time][day] = []
                            if ruleFollowed:
                                stack.append(
                                    (
                                        copySchedule,
                                        copyAvail,
                                        copyHoursDict,
                                        copyUnidealSet,
                                        copyAvailCount,
                                    )
                                )
                        if operationDone:
                            break
                    if operationDone:
                        break

        # Ensure that for each schedule each officer has 3 hours
        completeSolutions = []
        for solution, hourcount in solutions:
            noOptions = False
            queue = [None for _ in range(len(self.__Officers))]
            for officer in hourcount:
                queue[nameToObjectDict[officer].getOrder() - 1] = (
                    nameToObjectDict[officer],
                    (3 - hourcount[officer]),
                )
            for officer, count in queue:
                if count != 0:
                    extraHours = []
                    for time in range(len(solution)):
                        for day in range(len(solution[time])):
                            if (
                                type(solution[time][day]) == str
                                and officer.getSchedule()[time][day] != 0
                            ):
                                currVal = officer.getSchedule()[time][day]
                                if len(extraHours) < count:
                                    if len(extraHours) == 0:
                                        minLoc = 0
                                        minVal = currVal
                                    else:
                                        if currVal < minVal:
                                            minVal = currVal
                                            minLoc = len(extraHours)
                                    extraHours.append((time, day))
                                elif currVal > minVal:
                                    extraHours[minLoc] = (time, day)
                                    minVal = currVal
                                    for hour in range(len(extraHours)):
                                        checkVal = officer.getSchedule()[
                                            extraHours[hour][0]
                                        ][extraHours[hour][1]]
                                        if checkVal < minVal:
                                            minVal = checkVal
                                            minLoc = hour
                    if len(extraHours) == 0:
                        noOptions = True
                        break
                    for hour in extraHours:
                        current = solution[hour[0]][hour[1]]
                        solution[hour[0]][hour[1]] = (current, officer.getName())
            if not noOptions:
                completeSolutions.append(solution)
        return completeSolutions

    def compareSchedules(self, schedules, officers) -> list:
        # Comparison steps
        # 1. No Three Hours in one day
        # 2. Least Available but not preferred hours
        # 3. Fewest Joint Hours per person
        # 4. Least 2 Hours in 1 Day
        # 5. No 2 hours in a row
        # 6. Least Amount of Light Green
        # 7. Least Middle Hour of 3 Hour blocks
        # 8. Linear Fits ranking officers
        # Empty output list
        optimalSchedules = []

        # Baseline comparison values
        minAvailHours = 48
        minJointHours = 7  # Calculated by summing all non-zero joint hours for each person and subtracting 1 per person e.g. max case John 3 joint hrs Luca 3 joint hrs Alex 2 joint hrs 7 = 2 * (3 - 1) + (2 - 1)
        minTwoHours = 48
        minPrefHours = 48
        minMiddles = 38
        minIndivScore = 5

        # Dictionary for pulling officer object from name
        nameToObjectDict = {i.getName(): i for i in officers}

        for schedule in schedules:
            # Base state trackers for comparison criteria
            threeHours = False
            availHours = 0
            jointHours = 0
            twoHours = 0
            prefHours = 0
            middleHours = 0
            maxIndivScore = 0
            oneJoint = set()
            twoJoint = set()
            regressionData = {officer: 0 for officer in officers}

            # Iterates through schedules
            for day in range(5):
                # Trackers for how many hours in a day each officer has
                inDay = set()
                twoInDay = set()
                for time in range(timeOpen):
                    # Sets up joint-hour vs. solo-hour for correct data type
                    if type(schedule[time][day]) == tuple:
                        currentSlot = list(schedule[time][day])
                    else:
                        currentSlot = [schedule[time][day]]
                    # Iterates through officer(s) in slot
                    for officer in currentSlot:
                        # Checker for hours per day
                        if officer in twoInDay:
                            threeHours = True
                            break
                        elif officer in inDay:
                            twoInDay.add(officer)
                        else:
                            inDay.add(officer)
                        # Checks for joint hours per officer
                        if len(currentSlot) == 2:
                            if officer in oneJoint:
                                jointHours += 1
                                twoJoint.add(officer)
                            elif officer in twoJoint:
                                jointHours += 1
                            else:
                                oneJoint.add(officer)
                        # Checks what type hour is and adds to correct tracker also updates individual officer scores
                        if nameToObjectDict[officer].getSchedule()[time][day] == 1:
                            availHours += 1
                            regressionData[nameToObjectDict[officer]] += 3
                        elif nameToObjectDict[officer].getSchedule()[time][day] == 2:
                            prefHours += 1
                            regressionData[nameToObjectDict[officer]] += 1
                            # Checks if hours is in the middle of a block of three preferred hours
                            if (
                                time != 0
                                and time != 7
                                and (
                                    nameToObjectDict[officer].getSchedule()[time - 1][
                                        day
                                    ]
                                    == 2
                                    or nameToObjectDict[officer].getSchedule()[
                                        time - 1
                                    ][day]
                                    == 3
                                )
                                and (
                                    nameToObjectDict[officer].getSchedule()[time + 1][
                                        day
                                    ]
                                    == 2
                                    or nameToObjectDict[officer].getSchedule()[
                                        time + 1
                                    ][day]
                                    == 3
                                )
                            ):
                                middleHours += 1
                        if regressionData[nameToObjectDict[officer]] > maxIndivScore:
                            maxIndivScore = regressionData[nameToObjectDict[officer]]

                    if threeHours:
                        break
                # Increments counter for two hours in one day
                twoHours += len(twoInDay)
                if threeHours:
                    break
            if not threeHours:
                if availHours < minAvailHours:
                    optimalSchedules = []
                    optimalSchedules.append((schedule, regressionData))
                    minAvailHours = availHours
                    minJointHours = jointHours
                    minTwoHours = twoHours
                    minPrefHours = prefHours
                    minMiddles = middleHours
                    minIndivScore = maxIndivScore
                elif availHours == minAvailHours:
                    if jointHours < minJointHours:
                        optimalSchedules = []
                        optimalSchedules.append((schedule, regressionData))
                        minJointHours = jointHours
                        minTwoHours = twoHours
                        minPrefHours = prefHours
                        minMiddles = middleHours
                        minIndivScore = maxIndivScore
                    elif jointHours == minJointHours:
                        if twoHours < minTwoHours:
                            optimalSchedules = []
                            optimalSchedules.append((schedule, regressionData))
                            minTwoHours = twoHours
                            minPrefHours = prefHours
                            minMiddles = middleHours
                            minIndivScore = maxIndivScore
                        elif twoHours == minTwoHours:
                            if prefHours < minPrefHours:
                                optimalSchedules = []
                                optimalSchedules.append((schedule, regressionData))
                                minPrefHours = prefHours
                                minMiddles = middleHours
                                minIndivScore = maxIndivScore
                            elif prefHours == minPrefHours:
                                if middleHours < minMiddles:
                                    optimalSchedules = []
                                    optimalSchedules.append((schedule, regressionData))
                                    minMiddles = middleHours
                                    minIndivScore = maxIndivScore
                                elif middleHours == minMiddles:
                                    if maxIndivScore < minIndivScore:
                                        optimalSchedules = []
                                        optimalSchedules.append(
                                            (schedule, regressionData)
                                        )
                                        minIndivScore = maxIndivScore
                                    elif maxIndivScore == minIndivScore:
                                        optimalSchedules.append(
                                            (schedule, regressionData)
                                        )

        if len(optimalSchedules) > 1:
            # Linear Regression Comparison
            maxSSR = 5000
            # Post regression schedule(s)
            postRegress = []
            for schedule, data in optimalSchedules:
                n = len(officers)

                # For loop to traverse through all
                # element in an array
                for officerOne in range(n):
                    for officerTwo in range(0, n - officerOne - 1):
                        # Range of the array is from 0 to n-i-1
                        # Swap the elements if the element found
                        # is greater than the adjacent element
                        if data[officers[officerTwo]] > data[officers[officerTwo + 1]]:
                            officers[officerTwo], officers[officerTwo + 1] = (
                                officers[officerTwo + 1],
                                officers[officerTwo],
                            )
                        elif (
                            data[officers[officerTwo]] == data[officers[officerTwo + 1]]
                        ):
                            if (
                                officers[officerTwo].getOrder()
                                > officers[officerTwo + 1].getOrder()
                            ):
                                officers[officerTwo], officers[officerTwo + 1] = (
                                    officers[officerTwo + 1],
                                    officers[officerTwo],
                                )

                # Due to the points that are being compared we don't need to calculate everything for the R^2 value just the summed squared regression
                # Each schedule will have the same total sum of squares since all values are the same the orders just might be changed
                currSSR = 0

                for officer in range(len(officers)):
                    currSSR += ((officer + 1) - officers[officer].getOrder()) ** 2
                if currSSR < maxSSR:
                    maxSSR = currSSR
                    postRegress = []
                    postRegress.append(schedule)
                elif currSSR == maxSSR:
                    postRegress.append(schedule)
        else:
            return optimalSchedules[0]
        return postRegress

    def optimalSchedule(self) -> Schedule:
        legalSchedules = False
        allIdeal = True
        preferCycle = 0
        availCycle = 0
        while not legalSchedules:
            # Creates Next Set of Schedule Parameters to be checked
            if allIdeal:
                scheduleSet = []
                for i in self.__Officers:
                    scheduleSet.append(
                        Officer(
                            i.getName(),
                            i.getScheduleObject().idealOnly().hourSegments(),
                            i.getOrder(),
                            i.getReliability(),
                        )
                    )
                allIdeal = False
            elif preferCycle < len(self.__Officers):
                scheduleSet = []
                for i in self.__Officers:
                    if len(self.__Officers) - preferCycle <= i.getOrder():
                        scheduleSet.append(
                            Officer(
                                i.getName(),
                                i.getScheduleObject().preferredOnly().hourSegments(),
                                i.getOrder(),
                                i.getReliability(),
                            )
                        )
                    else:
                        scheduleSet.append(
                            Officer(
                                i.getName(),
                                i.getScheduleObject().idealOnly().hourSegments(),
                                i.getOrder(),
                                i.getReliability(),
                            )
                        )
                preferCycle += 1
            elif availCycle < len(self.__Officers):
                scheduleSet = []
                for i in self.__Officers:
                    if len(self.__Officers) - availCycle <= i.getOrder():
                        scheduleSet.append(
                            Officer(
                                i.getName(),
                                i.getScheduleObject().hourSegments(),
                                i.getOrder(),
                                i.getReliability(),
                            )
                        )
                    else:
                        scheduleSet.append(
                            Officer(
                                i.getName(),
                                i.getScheduleObject().preferredOnly().hourSegments(),
                                i.getOrder(),
                                i.getReliability(),
                            )
                        )
                availCycle += 1
            else:
                print("No possible schedules with one hour segments")
                return

            possibles = []

            if self.emptySlots(self.nameAvailability(scheduleSet)):
                availBools = [False for x in range(availCycle + preferCycle)]
                availBools[-1] = True
                while (
                    availBools != [True for x in range(len(self.__Officers))]
                    and possibles == []
                ):
                    scheduleSet = []
                    for i in range(len(availBools)):
                        if availBools[i]:
                            availBools[i] = False
                        else:
                            availBools[i] = True
                            break
                    for i in range(len(self.__Officers)):
                        if i < availCycle:
                            if availBools[
                                2 * len(self.__Officers) - self.__Officers[i].getOrder()
                            ]:
                                scheduleSet.append(
                                    Officer(
                                        self.__Officers[i].getName(),
                                        self.__Officers[i]
                                        .getScheduleObject()
                                        .hourSegments(),
                                        self.__Officers[i].getOrder(),
                                        self.__Officers[i].getReliability(),
                                    )
                                )
                            elif availBools[
                                len(self.__Officers) - self.__Officers[i].getOrder()
                            ]:
                                scheduleSet.append(
                                    Officer(
                                        self.__Officers[i].getName(),
                                        self.__Officers[i]
                                        .getScheduleObject()
                                        .preferredOnly()
                                        .hourSegments(),
                                        self.__Officers[i].getOrder(),
                                        self.__Officers[i].getReliability(),
                                    )
                                )
                            else:
                                scheduleSet.append(
                                    Officer(
                                        self.__Officers[i].getName(),
                                        self.__Officers[i]
                                        .getScheduleObject()
                                        .idealOnly()
                                        .hourSegments(),
                                        self.__Officers[i].getOrder(),
                                        self.__Officers[i].getReliability(),
                                    )
                                )
                        elif availBools[
                            len(self.__Officers) - self.__Officers[i].getOrder()
                        ]:
                            scheduleSet.append(
                                Officer(
                                    self.__Officers[i].getName(),
                                    self.__Officers[i]
                                    .getScheduleObject()
                                    .preferredOnly()
                                    .hourSegments(),
                                    self.__Officers[i].getOrder(),
                                    self.__Officers[i].getReliability(),
                                )
                            )
                        else:
                            scheduleSet.append(
                                Officer(
                                    self.__Officers[i].getName(),
                                    self.__Officers[i]
                                    .getScheduleObject()
                                    .idealOnly()
                                    .hourSegments(),
                                    self.__Officers[i].getOrder(),
                                    self.__Officers[i].getReliability(),
                                )
                            )

                    # Create set of possible schedules
                    possibles = self.createSchedules(scheduleSet)
                if possibles != []:
                    legalSchedules = True
        # Compare schedules using desired parameters
        return self.compareSchedules(possibles, scheduleSet)


def translateColor(rgbCode) -> int:
    rgbCode = str(rgbCode)
    if rgbCode == "FFFF0000":
        return 0
    elif rgbCode == "00000000" or rgbCode == "FFFFFFFF":
        return 1
    elif rgbCode == "FF00FF00":
        return 2
    elif rgbCode == "FF38761D":
        return 3
    else:
        print("Unknown Color:", rgbCode)
        return 6


def main():
    # User Settings
    numOfficers = 16
    officeOpen = 9
    officeClose = 5

    # Internal Settings
    if officeOpen >= officeClose:
        officeClose += 12
    global timeOpen
    timeOpen = officeClose - officeOpen
    timeIntervals = timeOpen * 2
    betweenRows = 2
    startRow = 6
    currentRow = startRow

    # Import spreadsheet
    filename = "Available"
    filename = filename + ".xlsx"
    availability = xl.load_workbook(filename, data_only=True)
    avail = availability["Officer Availability"]

    # Creation of list of officers
    officers = []
    print(
        "You will now be asked to input the order that each officer filled out their availability"
    )
    print(
        "Do this by entering a number 1-16, 1 being filled out first 16 being filled out last"
    )
    print(
        "Please be sure to enter each number only once if you enter one twice you will be asked to fix it"
    )
    numSet = set()
    for i in range(numOfficers):
        name = ""

        # Parsing sheet for needed data
        necessary = False
        while not necessary:
            # Get name
            if name == "":
                name = avail.cell(row=currentRow, column=1).value
                currentRow += betweenRows

            # Get schedule
            schedule = []
            for _ in range(timeIntervals):
                currentTime = []
                for day in range(5):
                    color = translateColor(
                        avail.cell(
                            row=currentRow, column=(day + 1) * 2
                        ).fill.start_color.index
                    )
                    if color == 6:
                        print(name, "messed up in cell:", currentRow, (day + 1))
                    currentTime.append(color)
                schedule.append(currentTime)
                currentRow += 1
            currentRow += 1
            necessary = True

        # Tentative for testing just puts priority in listed order from excel
        order = int(input(name + ": "))
        while order in numSet or order < 1 or order > numOfficers:
            print(
                "This number has either already been entered or is not between 0 and 16.\nPlease input a number that has yet to be input.\nThe following numbers have already been used\n"
                + str(numSet)
            )

            order = int(input(name + ": "))
        numSet.add(order)
        schedule = Schedule(schedule)

        officers.append(Officer(name, schedule, order))
    mastSched = MasterMaker(officers)
    finalSchedules = mastSched.optimalSchedule()

    # Predetermined list of colors to assign to each officer as their representation in final schedule
    finalColors = [
        "FFF56363",
        "FFFFABAB",
        "FFFFA97E",
        "FFFFE599",
        "FFF7FF36",
        "FFA9E178",
        "FF95B786",
        "FF63D297",
        "FFBBEAE7",
        "FF74A0E9",
        "FFA4C2F4",
        "FF8AA1C8",
        "FFB4A6FF",
        "FFD9D2E9",
        "FFD5A6BD",
        "FFDF61A3",
        "FFEF135B",
    ]
    nameToColorDict = {
        officers[i].getName(): finalColors[i] for i in range(len(officers))
    }

    # Creation of Excel sheet and visual chart for schedule
    output = xl.load_workbook("Template.xlsx")
    template = output.active
    sheetnum = 0

    for schedule in finalSchedules:
        currentSheet = output.copy_worksheet(template)
        sheetnum += 1
        currentSheet.title = "Schedule Option " + str(sheetnum)

        # Declares starting points for referencing cells
        startRow = 4
        startColumn = 2

        for time in range(len(schedule)):
            for day in range(len(schedule[time])):
                if type(schedule[time][day]) != tuple:
                    cell = currentSheet.cell(
                        row=(startRow + 2 * time), column=(startColumn + 2 * day)
                    )
                    cell.value = schedule[time][day]
                    cell.fill = PatternFill(
                        "solid", fgColor=nameToColorDict[schedule[time][day]]
                    )
                else:
                    currentSheet.unmerge_cells(
                        start_row=(startRow + 2 * time),
                        start_column=(startColumn + 2 * day),
                        end_row=(1 + startRow + 2 * time),
                        end_column=(1 + startColumn + 2 * day),
                    )
                    currentSheet.merge_cells(
                        start_row=(startRow + 2 * time),
                        start_column=(startColumn + 2 * day),
                        end_row=(1 + startRow + 2 * time),
                        end_column=(startColumn + 2 * day),
                    )
                    cellOne = currentSheet.cell(
                        row=(startRow + 2 * time), column=(startColumn + 2 * day)
                    )
                    cellOne.value = schedule[time][day][0]
                    cellOne.fill = PatternFill(
                        "solid", fgColor=nameToColorDict[schedule[time][day][0]]
                    )
                    currentSheet.merge_cells(
                        start_row=(startRow + 2 * time),
                        start_column=(1 + startColumn + 2 * day),
                        end_row=(1 + startRow + 2 * time),
                        end_column=(1 + startColumn + 2 * day),
                    )
                    cellTwo = currentSheet.cell(
                        row=(startRow + 2 * time), column=(1 + startColumn + 2 * day)
                    )
                    cellTwo._style = copy.copy(cellOne._style)
                    cellTwo.value = schedule[time][day][1]
                    cellTwo.fill = PatternFill(
                        "solid", fgColor=nameToColorDict[schedule[time][day][1]]
                    )

    output.save("OfficeHoursSchedule.xlsx")

    input("Press enter when done with program")


if __name__ == "__main__":
    main()
